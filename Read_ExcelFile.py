from openpyxl import load_workbook  

Workbook = load_workbook("example.xlsx")

sheet = Workbook.active 

print("Reading Excel File:")

for row in sheet.iter_rows(values_only=True):

    print(row)