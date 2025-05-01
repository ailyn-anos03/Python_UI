import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook


window = tk.Tk()
window.title("Student Score Tracker")
window.configure(bg="MistyRose3")


try:
    wb = load_workbook( "student_scores.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Score", "Result"]) 



# funtions
def show_data():
    data_window = tk.Toplevel(window)
    data_window.title("Student Scores")

    headers = ["Name", "Score", "Result"]
    for col, header in enumerate(headers):
        tk.Label(data_window, text=header, borderwidth=2, relief="groove", width=15, bg="MistyRose3", fg="black").grid(row=0, column=col, padx=5, pady=5)

 
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        for col_idx, value in enumerate(row):
            tk.Label(data_window, text=value, borderwidth=2, relief="groove", width=15).grid(row=row_idx, column=col_idx, padx=5, pady=5)

def average():
    total = 0
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None and row[0] != "Average": 
            try:
                score = float(row[1])
                total += score
                count += 1
            except ValueError:
                continue

    average = total / count if count > 0 else 0
    result = "Passed" if average >= 76 else "Failed"
    total_row = ["Average", f"{average:.2f}", result]

   
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if row[0] == "Average":
            ws.delete_rows(row_index)
            break


    ws.append(total_row)
    wb.save("student_scores.xlsx")

   
    tree.delete(*tree.get_children())
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

def add_data():
    name = name_entry.get().strip()
    score = score_entry.get().strip()

    if not name:
        messagebox.showerror("Input Error", "Name cannot be empty!")
        return

    try:
        score = int(score)
    except ValueError:
        messagebox.showerror("Input Error", "Score must be a numeric value!")
        return

    result = "Passed" if score >= 76 else "Failed"
    new_row = [name, score, result]

    # Insert the new data at the top
    ws.insert_rows(2)
    for col_index, value in enumerate(new_row, start=1):
        ws.cell(row=2, column=col_index, value=value)
    wb.save("student_scores.xlsx")
    messagebox.showinfo("Success", f"Added: {name}, Score: {score}, Result: {result}")
    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)

    
    tree.delete(*tree.get_children())
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

# labels
tk.Label(window, text="Student Name:", bg="MistyRose3", font="Times 12").grid(row=0, column=0, padx=10, pady=5, sticky="e")

tk.Label(window, text="Student Score:", bg="MistyRose3", font="Times 12").grid(row=1, column=0, padx=10, pady=5, sticky="e")

# table_headings
tree = ttk.Treeview(window, columns=("Name", "score", "Result"), show="headings")
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview.Heading", foreground="black", background="sienna4")

average()

# entries
score_entry = tk.Entry(window,font="Arial 10")
score_entry.grid(row=1, column=1, padx=10, pady=5)

name_entry = tk.Entry(window, font="Arial 10")
name_entry.grid(row=0, column=1, padx=10, pady=5)

# buttons
add_button = tk.Button(window, text="Add Data", command=add_data, bg="pink3", fg="white", font="Arial 10")
add_button.grid(row=2, column=0,  pady=10)

view_button = tk.Button(window, text="View Data", command=show_data, bg="pink3", fg="white", font="Arial 10")
view_button.grid(row=2, column=1, columnspan=2, pady=10)



window.mainloop()

