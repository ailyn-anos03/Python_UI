import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Initialize Excel file
file_path = "pos_data.xlsx"
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Item", "Quantity", "Price", "Total"])
    wb.save(file_path)

# Function to add item to Excel
def add_item():
    item = item_entry.get()
    quantity = quantity_entry.get()
    price = price_entry.get()

    if not item or not quantity or not price:
        messagebox.showerror("Error", "All fields are required!")
        return

    try:
        quantity = int(quantity)
        price = float(price)
        total = quantity * price

        wb = load_workbook(file_path)
        ws = wb.active
        ws.append([item, quantity, price, total])
        wb.save(file_path)

        messagebox.showinfo("Success", "Item added successfully!")
        item_entry.delete(0, tk.END)
        quantity_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)
    except ValueError:
        messagebox.showerror("Error", "Invalid input! Quantity must be an integer and Price must be a number.")

# Function to view all items
def view_items():
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        items_window = tk.Toplevel(root)
        items_window.title("Items List")

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for j, value in enumerate(row, start=1):
                tk.Label(items_window, text=value, borderwidth=1, relief="solid", width=15).grid(row=i, column=j)
    except Exception as e:
        messagebox.showerror("Error", f"Could not load items: {e}")

# GUI setup
root = tk.Tk()
root.title("POS System")

tk.Label(root, text="Item:").grid(row=0, column=0, padx=10, pady=5)
item_entry = tk.Entry(root)
item_entry.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Quantity:").grid(row=1, column=0, padx=10, pady=5)
quantity_entry = tk.Entry(root)
quantity_entry.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Price:").grid(row=2, column=0, padx=10, pady=5)
price_entry = tk.Entry(root)
price_entry.grid(row=2, column=1, padx=10, pady=5)

add_button = tk.Button(root, text="Add Item", command=add_item)
add_button.grid(row=3, column=0, columnspan=2, pady=10)

view_button = tk.Button(root, text="View Items", command=view_items)
view_button.grid(row=4, column=0, columnspan=2, pady=10)

root.mainloop()