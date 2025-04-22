from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Name"
ws["B1"] = "Age"
ws["A2"] = "Alice"
ws["B2"] = 25
ws["A3"] = "Bob"
ws["B3"] = 30


wb.save('example.xlsx')
print('Excel file created successfully')