from openpyxl import load_workbook  

Workbook = load_workbook("example.xlsx")

ws = Workbook.active 

ws["B2"] = 26  


Workbook.save('example.xlsx')

print("Excel file updated successfully!")