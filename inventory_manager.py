import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
from tkinter import messagebox, simpledialog
import os

EXCEL_FILE = 'inventory.xlsx'



def log_login_time(username):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        wb = load_workbook("inventory.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "LoginHistory"
        ws.append(["Username", "Timestamp"])
        wb.save("inventory.xlsx")

    if "LoginHistory" not in wb.sheetnames:
        login_ws = wb.create_sheet("LoginHistory")
        login_ws.append(["Username", "Timestamp"])
    else:
        login_ws = wb["LoginHistory"]

    login_ws.append([username, now])
    wb.save("inventory.xlsx")
    wb.close()

def init_credentials_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "RegisteredUsers"
        ws.append(["Username", "Password"])
        ws.append(["admin", "admin123"])
        wb.save(EXCEL_FILE)

def validate_login(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            wb.close()
            return True
    wb.close()
    return False

def user_exists(username):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            wb.close()
            return True
    wb.close()
    return False
# def show_register_page():
#     if 'login_frame' in globals() and 'register_frame' in globals():
#         login_frame.grid_forget()
#         register_frame.grid()
#     else:
#         messagebox.showerror("Error", "Frames are not initialized.")



    if not user or not pwd:
        messagebox.showwarning("Input Error", "Please fill in both fields.")
        return

    # Check if the user exists in the "RegisteredUsers" worksheet
    wb = load_workbook(EXCEL_FILE)
    if "RegisteredUsers" not in wb.sheetnames:
        ws = wb.create_sheet("RegisteredUsers")
        ws.append(["Username", "Password"])  # Add headers
        wb.save(EXCEL_FILE)
    else:
        ws = wb["RegisteredUsers"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user:
            wb.close()
            messagebox.showerror("Error", "Username already exists.")
            return

    # Register the user in the "RegisteredUsers" worksheet
    ws.append([user, pwd])
    wb.save(EXCEL_FILE)
    wb.close()

    messagebox.showinfo("Success", "User registered successfully!")
    show_login_page()

def register_user(username, password):
    wb = load_workbook(EXCEL_FILE)
    
    # Check if "RegisteredUsers" sheet exists, otherwise create it
    if "RegisteredUsers" not in wb.sheetnames:
        ws = wb.create_sheet("RegisteredUsers")
        ws.append(["Username", "Password"])
        
    else:
        ws = wb["RegisteredUsers"]
    
    ws.append([username, password])
    wb.save(EXCEL_FILE)
    wb.close()

# Initialize global variables
current_user = None

def login():
    global current_user
    user = login_username.get().strip()
    pwd = login_password.get().strip()

    if not user or not pwd:
        messagebox.showwarning("Input Error", "Please fill in both fields.")
        return

    # Validate login against the "RegisteredUsers" sheet
    try:
        wb = load_workbook(EXCEL_FILE)
        if "RegisteredUsers" in wb.sheetnames:
            ws = wb["RegisteredUsers"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == user and row[1] == pwd:
                    current_user = user
                    log_login_time(user)
                    messagebox.showinfo("Login Success", f"Welcome, {user}!")
                    inventoryWindow()
                    wb.close()
                    return
        wb.close()
    except FileNotFoundError:
        messagebox.showerror("Error", "Credentials file not found.")

    messagebox.showerror("Login Failed", "Invalid credentials.")

def show_login_page():
    if 'register_frame' in globals() and 'login_frame' in globals():
        login_frame.grid()
    else:
        messagebox.showerror("Error", "Frames are not initialized.")




def inventoryWindow():
    root.withdraw()
    PASSWORD = "admin123"
   
    def on_tab_changed(event):
        selected_tab = event.widget.index("current")
    
    
    # If user selects the protected tab (index 1)
        if selected_tab == 4 and not access_granted[0]:
            pwd = simpledialog.askstring("Password Required", "Enter password:", show="*")
            if pwd == PASSWORD:
                access_granted[0] = True
                

                def register():
                    user = reg_username.get().strip()
                    pwd = reg_password.get().strip()

                    if not user or not pwd:
                        messagebox.showwarning("Input Error", "Please fill in both fields.")
                        return

                    # Check if the user exists in the "RegisteredUsers" worksheet
                    wb = load_workbook(EXCEL_FILE)
                    if "RegisteredUsers" not in wb.sheetnames:
                        ws = wb.create_sheet("RegisteredUsers")
                        ws.append(["Username", "Password"])  # Add headers
                        wb.save(EXCEL_FILE)
                    else:
                        ws = wb["RegisteredUsers"]

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] == user:
                            wb.close()
                            messagebox.showerror("Error", "Username already exists.")
                            return True

                    # Register the user in the "RegisteredUsers" worksheet
                    ws.append([user, pwd])
                    wb.save(EXCEL_FILE)
                    wb.close()

                    messagebox.showinfo("Success", "User registered successfully!")

            
                
                # TAb 5

                register_frame = tk.Frame(tab5, bg="AntiqueWhite2")
                register_frame.grid(row=0, column=0, sticky="nsew")

                tk.Label(tab5, text="Register", font=("Arial", 16), bg="AntiqueWhite2").grid(row=0, column=0, columnspan=2, pady=15)
                tk.Label(tab5, text="Username", bg="AntiqueWhite2", font=("Arial", 12)).grid(row=1, column=0, padx=15, pady=10, sticky="w")
                reg_username = tk.Entry(tab5, font=("Arial", 12))
                reg_username.grid(row=1, column=1, padx=15, pady=10, sticky="we")
                tk.Label(tab5, text="Password", bg="AntiqueWhite2", font=("Arial", 12)).grid(row=2, column=0, padx=15, pady=10, sticky="w")
                reg_password = tk.Entry(tab5, show='*', font=("Arial", 12))
                reg_password.grid(row=2, column=1, padx=15, pady=10, sticky="we")

                # Transparent buttons
                tk.Button(tab5, text="Submit", command=register, font=("Arial", 12), bg="AntiqueWhite2", relief="ridge").grid(row=1, column=3, columnspan=2, pady=15)
                tk.Button(tab5, text="Back to Login", command=show_login_page, font=("Arial", 12), bg="AntiqueWhite2", relief="ridge").grid(row=2, column=3, columnspan=2)
                def add_item_user():
                    pas, user = reg_password.get(), reg_username.get()
                    if pas and user :
                            ws.append([pas, user])  # Append the new row at the end
                            wb.save("inventory.xlsx")
                            tree.insert("", "end", values=(pas, user))  # Insert at the end of the Treeview
                    
                        
                            clear_entries()
                            update_total()  # Update total after adding an item

                def view_data_user():
                    user_tree.delete(*user_tree.get_children())
                    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        user_tree.insert("", "end", values=row, iid=str(row_index))  

                def delete_item_user():
                    global selected_item_index
                    selected = user_tree.selection()
                    if selected:
                        row_index = int(selected[0])  # Row index
                        values = user_tree.item(selected[0], "values")
                        ws.delete_rows(row_index)  
                        wb.save("inventory.xlsx")
                        user_tree.delete(selected[0])  
                    
                        selected_item_index = None  # Reset after deletion
                        update_total()  # Update total after deleting an item

                def edit_item_user():
                    global selected_item_index
                    selected = user_tree.selection()
                    if selected:
                        selected_item_index = int(selected[0])  # Store row index
                        values = tree.item(selected[0], "values")
                        reg_password.delete(0, tk.END)
                        reg_password.insert(0, values[0])
                        reg_username.delete(0, tk.END)
                        reg_username.insert(0, values[1])
                        

                def update_item_user():
                    global selected_item_index
                    if selected_item_index:
                        new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
                        if all(new_values):
                            tree.item(str(selected_item_index), values=new_values) 
                            ws[selected_item_index][0].value = new_values[0]  
                            ws[selected_item_index][1].value = new_values[1]
                           
                            wb.save("inventory.xlsx")
                            
                            clear_entries()
                            selected_item_index = None 
                            update_total()  # Update total after updating an item

                def clear_entries_user():
                    reg_password.delete(0, tk.END)
                    reg_username.delete(0, tk.END)

                user_tree = ttk.Treeview(tab5, columns=("Username", "Password"), show="headings")
                style = ttk.Style()
                style.theme_use("clam")
                style.configure("Treeview.Heading", foreground="black", background="pink3")
                user_tree.grid(row=5, column=0, columnspan=5, sticky="nsew")

                tab5.grid_columnconfigure(0, weight=1)  
                tab5.grid_columnconfigure(1, weight=2)
                tab5.grid_columnconfigure(2, weight=1)
                tab5.grid_columnconfigure(3, weight=1)
                tab5.grid_columnconfigure(4, weight=1)
                tab5.grid_rowconfigure(7, weight=1)

                

                for col in ("Username", "Password"):
                    user_tree.heading(col, text=col)
                    user_tree.grid(row=5, column=0, columnspan=5)

                    try:
                        wb = load_workbook("inventory.xlsx")
                    except FileNotFoundError:
                        wb = Workbook()
                        ws = wb.active
                        ws.append("Username", "Password")  
                        wb.save("inventory.xlsx")

                        ws = wb.active
                        selected_item_index = None 
                
                
                
                tk.Button(tab5, text="Add", command=add_item_user, bg="AntiqueWhite2", relief="flat", font="Arial 10").grid(row=4, column=0, pady=5, sticky="we")
                tk.Button(tab5, text="Refresh", command=view_data_user, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=1, pady=5, sticky="we")
                tk.Button(tab5, text="Edit", command=edit_item_user, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=2, pady=5, sticky="we")
                tk.Button(tab5, text="Delete", command=delete_item_user, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=3, pady=5, sticky="we")
                tk.Button(tab5, text="Update", command=update_item_user, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=4, pady=5, sticky="we")

                view_data_user()  
            
            else:
                messagebox.showerror("Access Denied", "Incorrect password.")
                notebook.select(0)  # Revert to first tab

    access_granted = [False]
    
    Management = tk.Toplevel(root)
    Management.title("Inventory Management")
    Management.configure(bg="AntiqueWhite2")

    style = ttk.Style()
    style.configure('lefttab.TNotebook', tabposition='ws')  # Set tabs to the left side
    notebook = ttk.Notebook(Management, style='lefttab.TNotebook')

    # Add a second notebook for bottom tabs
    bottom_style = ttk.Style()
    bottom_style.configure('bottomtab.TNotebook', tabposition='s')  # Set tabs to the bottom
    bottom_notebook = ttk.Notebook(Management, style='bottomtab.TNotebook')

    tab1 = tk.Frame(notebook, bg="AntiqueWhite2")
    tab2 = tk.Frame(notebook, bg="AntiqueWhite2")
    tab3 = tk.Frame(notebook, bg="AntiqueWhite2")
    tab4 = tk.Frame(notebook, bg="AntiqueWhite2")
    tab5 = tk.Frame(notebook, bg="AntiqueWhite2")
                

    notebook.add(tab1, text="Home")
    notebook.add(tab2, text="Input")
    notebook.add(tab3, text="History")
    notebook.add(tab4, text="Inbox")
    notebook.add(tab5, text="User Management (ADMIN Only)")
    
    style.theme_use("clam")

    style.configure("TNotebook", background="RosyBrown2")  # Background of the notebook
    style.configure("TNotebook.Tab", background="AntiqueWhite3", foreground="black")
    style.map("TNotebook.Tab", background=[("selected", "AntiqueWhite2")], foreground=[("selected", "black")])


    notebook.grid(sticky="nsew")
    style.configure('lefttab.TNotebook', tabposition='wn')  # Set tabs to the left side
    notebook.bind("<<NotebookTabChanged>>", on_tab_changed)

    # Make the notebook stretchable
    Management.grid_columnconfigure(0, weight=1)
    Management.grid_rowconfigure(0, weight=1)

    try:
        wb = load_workbook("inventory.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Item", "Quantity", "Price"])  
        wb.save("inventory.xlsx")

    ws = wb.active
    selected_item_index = None 
    def add_item():
        item, quantity, price = entry_item.get(), entry_quantity.get(), entry_price.get()
        if item and quantity and price:
            ws.append([item, quantity, price])  # Append the new row at the end
            wb.save("inventory.xlsx")
            tree.insert("", "end", values=(item, quantity, price))  # Insert at the end of the Treeview
            log_action("Add", item, quantity, price)  # Log action to history
            log_action_to_inbox("Add", item, quantity, price)  # Log action to inbox
            clear_entries()
            update_total()  # Update total after adding an item

    def view_data():
        tree.delete(*tree.get_children())
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            tree.insert("", "end", values=row, iid=str(row_index))  

    def delete_item():
        global selected_item_index
        selected = tree.selection()
        if selected:
            row_index = int(selected[0])  # Row index
            values = tree.item(selected[0], "values")
            ws.delete_rows(row_index)  
            wb.save("inventory.xlsx")
            tree.delete(selected[0])  
            log_action("Delete", values[0], values[1], values[2])  # Log action to history
            log_action_to_inbox("Delete", values[0], values[1], values[2])  # Log action to inbox
            selected_item_index = None  # Reset after deletion
            update_total()  # Update total after deleting an item

    def edit_item():
        global selected_item_index
        selected = tree.selection()
        if selected:
            selected_item_index = int(selected[0])  # Store row index
            values = tree.item(selected[0], "values")
            entry_item.delete(0, tk.END)
            entry_item.insert(0, values[0])
            entry_quantity.delete(0, tk.END)
            entry_quantity.insert(0, values[1])
            entry_price.delete(0, tk.END)
            entry_price.insert(0, values[2])

    def update_item():
        global selected_item_index
        if selected_item_index:
            new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
            if all(new_values):
                tree.item(str(selected_item_index), values=new_values) 
                ws[selected_item_index][0].value = new_values[0]  
                ws[selected_item_index][1].value = new_values[1]
                ws[selected_item_index][2].value = new_values[2]
                wb.save("inventory.xlsx")
                log_action("Update", new_values[0], new_values[1], new_values[2])  # Log action to history
                log_action_to_inbox("Update", new_values[0], new_values[1], new_values[2])  # Log action to inbox
                clear_entries()
                selected_item_index = None 
                update_total()  # Update total after updating an item

    def clear_entries():
        entry_item.delete(0, tk.END)
        entry_quantity.delete(0, tk.END)
        entry_price.delete(0, tk.END)

    def logout():
        if Management.winfo_exists():  # Check if the window exists
            Management.destroy()
        root.deiconify()
    Management.protocol("WM_DELETE_WINDOW", logout)

    label1 = tk.Label(tab1, text="Inventory Management", bg="AntiqueWhite2", font="Arial 14 bold")
    label1.grid(row=0, column=0, padx=10, pady=10)

    tk.Button(tab1, text="Logout",command=logout, font="Arial 10").grid(row=0, column=5, pady=5, sticky="w")
   
   

   
 
    #Tab 2 - Input

    tk.Label(tab2, text="Item", bg="AntiqueWhite2", font="Calibri 11").grid(row=0, column=0, sticky="w")
    tk.Label(tab2, text="Quantity", bg="AntiqueWhite2", font="Calibri 11").grid(row=1, column=0, sticky="w")
    tk.Label(tab2, text="Price", bg="AntiqueWhite2", font="Calibri 11").grid(row=2, column=0, sticky="w")

    entry_item, entry_quantity, entry_price = tk.Entry(tab2, font="Times 11", justify="left"), tk.Entry(tab2, font="Times 11", justify="left"), tk.Entry(tab2, font="Times 11", justify="left")

    entry_item.grid(row=0, column=1, sticky="we")
    entry_quantity.grid(row=1, column=1,  sticky="we")
    entry_price.grid(row=2, column=1, sticky="we")

    tk.Button(tab2, text="Add", command=add_item, bg="AntiqueWhite2", relief="flat", font="Arial 10").grid(row=4, column=0, pady=5, sticky="we")
    tk.Button(tab2, text="Refresh", command=view_data, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=1, pady=5, sticky="we")
    tk.Button(tab2, text="Edit", command=edit_item, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=2, pady=5, sticky="we")
    tk.Button(tab2, text="Delete", command=delete_item, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=3, pady=5, sticky="we")
    tk.Button(tab2, text="Update", command=update_item, bg="AntiqueWhite2", relief="flat",font="Arial 10").grid(row=4, column=4, pady=5, sticky="we")

    tree = ttk.Treeview(tab2, columns=("Item", "Quantity", "Price"), show="headings")
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview.Heading", foreground="black", background="pink3")
    tree.grid(row=7, column=0, columnspan=5, sticky="nsew")



    tab2.grid_columnconfigure(0, weight=1)  
    tab2.grid_columnconfigure(1, weight=2)
    tab2.grid_columnconfigure(2, weight=1)
    tab2.grid_columnconfigure(3, weight=1)
    tab2.grid_columnconfigure(4, weight=1)
    tab2.grid_rowconfigure(7, weight=1)

    for col in ("Item", "Quantity", "Price"):
        tree.heading(col, text=col)
    tree.grid(row=7, column=0, columnspan=5)

    def search(event):
        query = entry.get().lower()
        tree.delete(*tree.get_children())  

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if query in str(row[0]).lower(): 
                tree.insert("", "end", values=row, iid=str(row_index))

    label = tk.Label(tab2, text="Search", bg="AntiqueWhite2", font="Times 11")
    label.grid(row=6, column=0, pady=10, sticky="w")
    entry = tk.Entry(tab2, font="Times 11", justify="left")
    entry.grid(row=6, column=1, columnspan=4, pady=5, sticky="we")
    entry.bind("<KeyRelease>", search)

    tab2.grid_columnconfigure(1, weight=1)  
    view_data()

    


    def update_total():
        def deduct_quantity():
            selected = tree.selection()
            if selected:
                values = tree.item(selected[0], "values")
                item_name = values[0]
                current_quantity = values[1]

                try:
                    current_quantity = int(current_quantity)
                    deduction = int(spinbox_deduction.get())
                    if deduction > current_quantity:
                        messagebox.showerror("Error", "Deduction exceeds current quantity.")
                        return

                    new_quantity = current_quantity - deduction
                    ws[int(selected[0])][1].value = new_quantity  # Update quantity in Excel
                    wb.save("inventory.xlsx")
                    tree.item(selected[0], values=(item_name, new_quantity, values[2]))  # Update Treeview
                    update_total()  # Update total after deduction
                    messagebox.showinfo("Success", f"Deducted {deduction} from {item_name}.")
                    
                    # Log action to history and inbox
                    log_action("Deduct", item_name, deduction, values[2])
                    log_action_to_inbox("Deduct", item_name, deduction, values[2])
                except ValueError:
                    messagebox.showerror("Error", "Invalid quantity or deduction value.")
            else:
                messagebox.showwarning("Warning", "No item selected.")

        tk.Label(tab2, text="Deduct Quantity", bg="AntiqueWhite2", font="Calibri 11").grid(row=5, column=0, sticky="w")
        spinbox_deduction = tk.Spinbox(tab2, from_=0, to=1000, increment=1, font="Times 11", justify="left")
        spinbox_deduction.grid(row=5, column=1, sticky="we")
        tk.Button(tab2, text="Deduct", command=deduct_quantity, bg="AntiqueWhite2", relief="flat", font="Arial 10").grid(row=5, column=2, pady=5, sticky="we")
        total = 0
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] is not None and row[2] is not None:  
                try:
                    quantity = float(row[1])
                    price = float(row[2])
                    total += quantity * price
                except ValueError:
                    continue

        total_row = ["Total", "", f"{total:.2f}"]

        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            if row[0] == "Total":
                ws.delete_rows(row_index)
                break

        ws.append(total_row)
        wb.save("inventory.xlsx")

    
        tree.delete(*tree.get_children())
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            tree.insert("", "end", values=row, iid=str(row_index))

    
        view_data()

    label2 = tk.Label(tab3, text="History", bg="AntiqueWhite2", font="Arial 14 bold", justify="center")
    label2.grid(row=0, column=0, padx=10, pady=10)

    history_tree = ttk.Treeview(tab3, columns=("Action", "Item", "Quantity", "Price", "Timestamp"), show="headings")
    history_tree.grid(row=1, column=0, columnspan=5, sticky="nsew")

    for col in ("Action", "Item", "Quantity", "Price", "Timestamp"):
        history_tree.heading(col, text=col)

    tab3.grid_columnconfigure(0, weight=1)
    tab3.grid_rowconfigure(1, weight=1)

    # Create a new worksheet for history if it doesn't exist
    if "History" not in wb.sheetnames:
        history_ws = wb.create_sheet("History")
        history_ws.append(["Action", "Item", "Quantity", "Price", "Timestamp"])
    else:
        history_ws = wb["History"]

    def log_action(action, item, quantity, price):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history_ws.append([action, item, quantity, price, timestamp])
        wb.save("inventory.xlsx")
        history_tree.insert("", "end", values=(action, item, quantity, price, timestamp))
        messagebox.showinfo("Action Logged", f"{action} action performed on item '{item}'.")

    def add_item_with_logging():
        item, quantity, price = entry_item.get(), entry_quantity.get(), entry_price.get()
        add_item()  # Call the original add_item function
        if item and quantity and price:
            log_action("Add", item, quantity, price)

    def delete_item_with_logging():
        global selected_item_index
        selected = tree.selection()
        if selected:
            values = tree.item(selected[0], "values")
            delete_item()  # Call the original delete_item function
            log_action("Delete", values[0], values[1], values[2])

    def update_item_with_logging():
        global selected_item_index
        if selected_item_index:
            new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
            update_item()  # Call the original update_item function
            if all(new_values):
                log_action("Update", new_values[0], new_values[1], new_values[2])
    
    tk.Button(tab2, text="Add", command=add_item_with_logging, bg="AntiqueWhite2", relief="flat", font="Arial 10").grid(row=4, column=0, pady=5, sticky="we")
    tk.Button(tab2, text="Delete", command=delete_item_with_logging, bg="AntiqueWhite2", relief="flat", font="Arial 10").grid(row=4, column=3, pady=5, sticky="we")
    tk.Button(tab2, text="Update", command=update_item_with_logging, bg="AntiqueWhite2", relief="flat", font="Arial 10").grid(row=4, column=4, pady=5, sticky="we")

    for row in history_ws.iter_rows(min_row=2, values_only=True):
        history_tree.insert("", "end", values=row)
    
    def search_history(event):
        query = history_entry.get().lower()
        history_tree.delete(*history_tree.get_children())  

        for row_index, row in enumerate(history_ws.iter_rows(min_row=2, values_only=True), start=2):
            if query in " ".join(map(str, row)).lower(): 
                history_tree.insert("", "end", values=row, iid=str(row_index))

    history_label = tk.Label(tab3, text="Search History", bg="AntiqueWhite2", font="Times 11")
    history_label.grid(row=2, column=0, pady=10, sticky="w")
    history_entry = tk.Entry(tab3, font="Times 11", justify="left")
    history_entry.grid(row=2, column=1, columnspan=4, pady=5, sticky="we")
    history_entry.bind("<KeyRelease>", search_history)

    tab3.grid_columnconfigure(1, weight=1)

    view_data()
    update_total()

    # Tab 4 - Inbox
    label3 = tk.Label(tab4, text="Inbox", bg="AntiqueWhite2", font="Arial 14 bold")
    label3.grid(row=0, column=0, padx=10, pady=10)

    inbox_tree = ttk.Treeview(tab4, columns=("Sender", "Message", "Timestamp"), show="headings")
    inbox_tree.grid(row=1, column=0, columnspan=5, sticky="nsew")

    for col in ("Sender", "Message", "Timestamp"):
        inbox_tree.heading(col, text=col)

    # Create a new worksheet for Inbox if it doesn't exist
    if "Inbox" not in wb.sheetnames:
        inbox_ws = wb.create_sheet("Inbox")
        inbox_ws.append(["Sender", "Message", "Timestamp"])
        wb.save("inventory.xlsx")
    else:
        inbox_ws = wb["Inbox"]

    # Function to display messages in the inbox treeview
    def display_inbox_message(sender, message, timestamp):
        inbox_tree.insert("", "end", values=(sender, message, timestamp))

    # Load initial inbox data
    def load_inbox_data():
        inbox_tree.delete(*inbox_tree.get_children())
        for row in inbox_ws.iter_rows(min_row=2, values_only=True):
            display_inbox_message(row[0], row[1], row[2])

    load_inbox_data()

    tab4.grid_columnconfigure(0, weight=1)
    tab4.grid_rowconfigure(1, weight=1)

    # Function to log messages to the Inbox worksheet
    def log_inbox_message(sender, message):
        global current_user
        sender = current_user if current_user else "System"
        timestamp = datetime.now().strftime("%d %b %Y, %I:%M %p")  # User-friendly timestamp format
        inbox_ws.append([sender, message, timestamp])
        wb.save("inventory.xlsx")
        display_inbox_message(sender, message, timestamp)

    # Automatically log messages to Inbox based on history actions
    def log_action_to_inbox(action, item, quantity, price):
        message = f"{action} performed on item '{item}' with quantity '{quantity}' and price '{price}'."
        log_inbox_message("System", message)

    # Update the history logging functions to also log to the Inbox
    def add_item_with_logging():
        item, quantity, price = entry_item.get(), entry_quantity.get(), entry_price.get()
        add_item()  # Call the original add_item function
        if item and quantity and price:
            log_action("Add", item, quantity, price)
            log_action_to_inbox("Add", item, quantity, price)

    def delete_item_with_logging():
        global selected_item_index
        selected = tree.selection()
        if selected:
            values = tree.item(selected[0], "values")
            delete_item()  # Call the original delete_item function
            log_action("Delete", values[0], values[1], values[2])
            log_action_to_inbox("Delete", values[0], values[1], values[2])

    def update_item_with_logging():
        global selected_item_index
        if selected_item_index:
            new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
            update_item()  # Call the original update_item function
            if all(new_values):
                log_action("Update", new_values[0], new_values[1], new_values[2])
                log_action_to_inbox("Update", new_values[0], new_values[1], new_values[2])

    # Automatically update the Inbox when a new entry is added to the History worksheet
    def sync_inbox_with_history():
        history_rows = list(history_ws.iter_rows(min_row=2, values_only=True))
        inbox_rows = list(inbox_ws.iter_rows(min_row=2, values_only=True))

        # Add missing history rows to the Inbox
        for row in history_rows[len(inbox_rows):]:
            action, item, quantity, price, timestamp = row
            log_action_to_inbox(action, item, quantity, price)

    # Call sync function whenever the history is updated
    sync_inbox_with_history()

    # Search functionality for the Inbox
    def search_inbox():
        query = inbox_entry.get().lower()
        inbox_tree.delete(*inbox_tree.get_children())

        for row in inbox_ws.iter_rows(min_row=2, values_only=True):
            if query in " ".join(map(str, row)).lower():
                inbox_tree.insert("", "end", values=row)

    inbox_label = tk.Label(tab4, text="Search Inbox", bg="AntiqueWhite2", font="Times 11")
    inbox_label.grid(row=2, column=0, pady=10, sticky="w")
    inbox_entry = tk.Entry(tab4, font="Times 11", justify="left")
    inbox_entry.grid(row=2, column=1, columnspan=4, pady=5, sticky="we")
    inbox_entry.bind("<KeyRelease>", lambda _: search_inbox())

    
    
   






    
   
    
    
init_credentials_file()

root = tk.Tk()
root.title("Login & Admin Management")

root.configure(bg="AntiqueWhite2")
root.resizable(False, False)

# ---------------------- Login Frame ---------------------

login_frame = tk.Frame(root, bg="AntiqueWhite2")
login_frame.grid(row=0, column=0, sticky="nsew")

tk.Label(login_frame, text="Login", font=("Arial", 16), bg="AntiqueWhite2").grid(row=0, column=0, columnspan=2, pady=15)

def on_entry_click(event, entry, placeholder):
    if entry.get() == placeholder:
        entry.delete(0, tk.END)
        entry.config(fg="black")

def on_focusout(event, entry, placeholder):
    if entry.get() == "":
        entry.insert(0, placeholder)
        entry.config(fg="grey")

login_username = tk.Entry(login_frame, font=("Arial", 12), fg="grey")
login_username.insert(0, "Username")
login_username.bind("<FocusIn>", lambda event: on_entry_click(event, login_username, "Username"))
login_username.bind("<FocusOut>", lambda event: on_focusout(event, login_username, "Username"))
login_username.grid(row=1, column=0, columnspan=2, padx=15, pady=10, sticky="we")

login_password = tk.Entry(login_frame, font=("Arial", 12), fg="grey", show="")
login_password.insert(0, "Password")
login_password.bind("<FocusIn>", lambda event: on_entry_click(event, login_password, "Password"))
login_password.bind("<FocusOut>", lambda event: on_focusout(event, login_password, "Password"))
login_password.grid(row=2, column=0, columnspan=2, padx=15, pady=10, sticky="we")

# Transparent buttons
tk.Button(login_frame, text="Login", command=login, font=("Arial", 12), bg="AntiqueWhite2", relief="flat").grid(row=3, column=0, columnspan=2, pady=15)
# tk.Button(login_frame, text="Register", command=show_register_page, font=("Arial", 12), bg="AntiqueWhite2", relief="flat").grid(row=4, column=0, columnspan=2)


root.mainloop()
